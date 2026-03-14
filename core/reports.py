# =============================================================================
#   core/reports.py — VaultSentry v1.0
#   Egyan | Red Parrot Accounting Ltd
#
#   Excel report generator — three sheets: Scan Results, Summary, Alert History.
# =============================================================================

import os
import shutil
import sqlite3
from datetime import datetime

from config import DB_PATH, REPORT_DIR, APP_NAME, APP_VERSION
from logger import log

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def generate_excel_report(results=None) -> str | None:
    """
    Generate a colour-coded Excel report.
    Saved to REPORT_DIR and also copied to Downloads.
    Returns the report path, or None if openpyxl is not installed.
    """
    if not EXCEL_AVAILABLE:
        log.warning(
            "openpyxl not installed — Excel report skipped. "
            "Install with: pip install openpyxl"
        )
        return None

    os.makedirs(REPORT_DIR, exist_ok=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_name = f"VaultSentry_Report_{timestamp}.xlsx"
    report_path = os.path.join(REPORT_DIR, report_name)

    wb = openpyxl.Workbook()

    # Shared styles
    header_fill   = PatternFill("solid", fgColor="1F3864")
    ok_fill       = PatternFill("solid", fgColor="C6EFCE")
    warning_fill  = PatternFill("solid", fgColor="FFEB9C")
    critical_fill = PatternFill("solid", fgColor="FFC7CE")
    info_fill     = PatternFill("solid", fgColor="DDEBF7")
    title_fill    = PatternFill("solid", fgColor="1a3a5c")
    subtitle_fill = PatternFill("solid", fgColor="BDD7EE")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, color="FFFFFF", size=14)
    bold_font   = Font(bold=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    severity_fills = {
        "CRITICAL": critical_fill,
        "WARNING":  warning_fill,
        "OK":       ok_fill,
        "INFO":     info_fill,
    }

    subtitle_text = (
        f"Red Parrot Accounting Ltd | "
        f"{datetime.now().strftime('%d %B %Y, %H:%M')}"
    )

    # Sheet 1: Scan Results
    ws1       = wb.active
    ws1.title = "Scan Results"
    _write_title_row(ws1, "A1:F1",
                     f"{APP_NAME} v{APP_VERSION} — Backup Integrity Scan Report",
                     title_fill, title_font)
    _write_subtitle_row(ws1, "A2:F2", subtitle_text, subtitle_fill)

    headers = ["#", "File Path", "Status", "Severity", "Details", "Scan Time"]
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[3].height = 20
    ws1.freeze_panes = "A4"

    scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not results:
        results = _load_results_from_db()

    for row_num, result in enumerate(results, start=4):
        severity = result.get("severity", "INFO")
        row_fill = severity_fills.get(severity, info_fill)
        values   = [
            row_num - 3,
            result.get("filepath", ""),
            result.get("status", ""),
            severity,
            result.get("detail", "")[:200],
            scan_time,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws1.cell(row=row_num, column=col, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 60
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 60
    ws1.column_dimensions["F"].width = 22

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    _write_title_row(ws2, "A1:C1", f"{APP_NAME} — Summary Dashboard",
                     title_fill, title_font)
    _write_subtitle_row(ws2, "A2:C2",
                        f"Red Parrot Accounting Ltd | {datetime.now().strftime('%d %B %Y')}",
                        subtitle_fill)

    stats    = _load_summary_stats()
    last_run = stats["last_run"]

    summary_rows = [
        ("",               "",                   ""),
        ("Metric",         "Count",              "Status"),
        ("Files OK",       stats["ok"],
         "✓ Good" if stats["ok"] > 0 else "—"),
        ("Files Changed",  stats["changed"],
         "⚠ Review" if stats["changed"] > 0 else "✓ None"),
        ("Files Missing",  stats["missing"],
         "✗ Critical" if stats["missing"] > 0 else "✓ None"),
        ("Corrupt Files",  stats["corrupt"],
         "✗ Critical" if stats["corrupt"] > 0 else "✓ None"),
        ("New Files",      stats["new"],          "ℹ Check"),
        ("",               "",                   ""),
        ("Critical Alerts (all time)", stats["crit_alerts"], ""),
        ("Warning Alerts (all time)",  stats["warn_alerts"], ""),
    ]
    if last_run:
        summary_rows += [
            ("",              "",               ""),
            ("Last Scan",     last_run[0],      ""),
            ("Files Scanned", last_run[1],      ""),
            ("Scan Duration", f"{last_run[2]}s", ""),
        ]

    for row_num, (label, value, status) in enumerate(summary_rows, start=3):
        if label == "Metric":
            for col, val in enumerate([label, value, status], start=1):
                c = ws2.cell(row=row_num, column=col, value=val)
                c.fill = header_fill
                c.font = header_font
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            continue
        ws2.cell(row=row_num, column=1, value=label).font = bold_font
        ws2.cell(row=row_num, column=2, value=value)
        status_cell = ws2.cell(row=row_num, column=3, value=status)
        if "✗" in str(status) or "Critical" in str(status):
            status_cell.fill = critical_fill
        elif "⚠" in str(status) or "Review" in str(status):
            status_cell.fill = warning_fill
        elif "✓" in str(status) or "Good" in str(status):
            status_cell.fill = ok_fill
        for col in range(1, 4):
            ws2.cell(row=row_num, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 18

    # Sheet 3: Alert History
    ws3 = wb.create_sheet("Alert History")
    _write_title_row(ws3, "A1:F1", f"{APP_NAME} — All-Time Alert History",
                     title_fill, title_font)
    for col, h in enumerate(
        ["#", "Timestamp", "Severity", "Alert Type", "File", "Details"], start=1
    ):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(
        "SELECT timestamp, severity, alert_type, filepath, details "
        "FROM alerts ORDER BY id DESC LIMIT 500"
    )
    alert_rows = cur.fetchall()
    conn.close()

    for i, (ts, sev, atype, fp, det) in enumerate(alert_rows, start=1):
        row_fill = (
            critical_fill if sev == "CRITICAL"
            else warning_fill if sev == "WARNING"
            else info_fill
        )
        for col, val in enumerate(
            [i, ts, sev, atype, fp, str(det)[:150]], start=1
        ):
            cell = ws3.cell(row=i + 2, column=col, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 35
    ws3.column_dimensions["E"].width = 55
    ws3.column_dimensions["F"].width = 50
    ws3.freeze_panes = "A3"

    wb.save(report_path)
    log.info("Excel report saved: %s", report_path)

    downloads_path = os.path.join(
        os.path.expanduser("~"), "Downloads",
        f"VaultSentry_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    try:
        shutil.copy2(report_path, downloads_path)
        log.info("Report also copied to Downloads: %s", downloads_path)
    except Exception:
        pass

    return report_path


# =============================================================================
#   Private helpers
# =============================================================================

def _write_title_row(ws, cell_range, text, fill, font, row_height=28):
    ws.merge_cells(cell_range)
    first_cell       = ws[cell_range.split(":")[0]]
    first_cell.value = text
    first_cell.fill  = fill
    first_cell.font  = font
    first_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = row_height


def _write_subtitle_row(ws, cell_range, text, fill):
    ws.merge_cells(cell_range)
    first_cell           = ws[cell_range.split(":")[0]]
    first_cell.value     = text
    first_cell.fill      = fill
    first_cell.font      = Font(italic=True, size=10)
    first_cell.alignment = Alignment(horizontal="center")


def _load_results_from_db():
    if not os.path.exists(DB_PATH):
        return []
    conn    = sqlite3.connect(DB_PATH)
    cur     = conn.cursor()
    cur.execute("SELECT filepath, status FROM file_hashes ORDER BY status DESC")
    db_rows = cur.fetchall()
    conn.close()
    return [
        {"filepath": r[0], "status": r[1], "detail": "",
         "severity": "OK" if r[1] == "OK" else "WARNING"}
        for r in db_rows
    ]


def _load_summary_stats():
    if not os.path.exists(DB_PATH):
        return {"ok": 0, "changed": 0, "missing": 0, "corrupt": 0, "new": 0,
                "crit_alerts": 0, "warn_alerts": 0, "last_run": None}
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    def count(where):
        cur.execute(f"SELECT COUNT(*) FROM file_hashes WHERE {where}")
        return cur.fetchone()[0]

    stats = {
        "ok":      count("status='OK'"),
        "changed": count("status='CHANGED'"),
        "missing": count("status='MISSING'"),
        "corrupt": count("status='CORRUPT'"),
        "new":     count("status='NEW'"),
    }
    cur.execute("SELECT COUNT(*) FROM alerts WHERE severity='CRITICAL'")
    stats["crit_alerts"] = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM alerts WHERE severity='WARNING'")
    stats["warn_alerts"] = cur.fetchone()[0]
    cur.execute(
        "SELECT run_time, files_scanned, duration_secs "
        "FROM scan_runs ORDER BY id DESC LIMIT 1"
    )
    stats["last_run"] = cur.fetchone()
    conn.close()
    return stats
